from openpyxl import Workbook
wb = Workbook()  # 使用Workbook()，並存為wb變數
                 # 以利後續使用wb變數來呼叫各種function來使用

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 26
ws['A2'] = 'Brian'

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A4'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")